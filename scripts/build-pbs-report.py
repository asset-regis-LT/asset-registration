#!/usr/bin/env python3
"""Phase 3: the clean PBS reporting list, built from the backfilled canonical
fields on the data branch. Level 1 area/sistem, Level 2 sub-sistem, Level 3
component - one row per canonical PBS, no unit/typo variants.

  git --work-tree=/tmp/dco checkout origin/data -- data/inspections
  git reset -q -- data/inspections
  python3 scripts/build-pbs-report.py [/tmp/dco/data/inspections]
"""
import json
import glob
import sys
import collections
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REPO = Path(__file__).resolve().parent.parent
INSPECTIONS = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/dco/data/inspections")
MASTER = REPO / "PBS_MASTER_CANONICAL.xlsx"
OUT = REPO / "PBS_STRUKTUR_CANONICAL.xlsx"

# canonical master -> names by code
ws = openpyxl.load_workbook(MASTER, data_only=True).worksheets[0]
m_l1, m_l2, m_l3 = {}, {}, {}
for kode, lvl, jenis, komp in list(ws.iter_rows(values_only=True))[2:]:
    if lvl == 1: m_l1[str(kode)] = komp
    elif lvl == 2: m_l2[str(kode)] = komp
    elif lvl == 3: m_l3[str(kode)] = komp

def pbskey(c):
    return [int(x) for x in str(c).split(".")]

# aggregate records by canonical PBS
insp = collections.Counter()                 # cpbs -> record count
units = collections.defaultdict(set)         # cpbs -> {unitLabel}
name = {}                                     # cpbs -> namaAsetCanonical (from records)
sub = {}                                      # cpbs -> subsistemCanonical (from records)
for p in glob.glob(str(INSPECTIONS / "*.json")):
    d = json.load(open(p, encoding="utf-8"))
    c = (d.get("pbsCanonical") or d.get("nomorPBS") or "").strip()
    if c.count(".") != 2:
        continue
    insp[c] += 1
    u = (d.get("unitLabel") or "").strip()
    if u:
        units[c].add(u)
    name.setdefault(c, (d.get("namaAsetCanonical") or d.get("namaAset") or "").strip())
    sub.setdefault(c, (d.get("subsistemCanonical") or d.get("subsistem") or "").strip())

codes = sorted(insp, key=pbskey)

rows = []  # (kode, level, area, subsis, komp, n_insp, n_unit)
seen_sys, seen_mid = set(), set()
for c in codes:
    sysn = c.split(".")[0]
    mid = ".".join(c.split(".")[:2])
    if sysn not in seen_sys:
        seen_sys.add(sysn)
        tot = sum(insp[x] for x in codes if x.split(".")[0] == sysn)
        rows.append((sysn, 1, m_l1.get(sysn, f"Sistem {sysn}"), "", "", tot, ""))
    if mid not in seen_mid:
        seen_mid.add(mid)
        tot = sum(insp[x] for x in codes if ".".join(x.split(".")[:2]) == mid)
        rows.append((mid, 2, "", m_l2.get(mid, sub.get(c, "")), "", tot, ""))
    rows.append((c, 3, "", "", m_l3.get(c) or name.get(c, ""), insp[c], len(units[c])))

wb = openpyxl.Workbook()
sh = wb.active
sh.title = "Struktur PBS Kanonik"
sh.append(["Kode PBS", "Level", "Area / Sistem", "Sub-sistem",
           "Komponen / Peralatan", "Jumlah Inspeksi", "Jumlah Unit"])
for cell in sh[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F4F4F")

f1 = PatternFill("solid", fgColor="D9E2F3")
f2 = PatternFill("solid", fgColor="EDEDED")
for kode, lvl, area, subs, komp, ni, nu in rows:
    sh.append([kode, lvl, area, subs, komp, ni, nu])
    r = sh[sh.max_row]
    if lvl == 1:
        for cc in r: cc.font = Font(bold=True); cc.fill = f1
    elif lvl == 2:
        for cc in r: cc.font = Font(bold=True); cc.fill = f2
        r[3].alignment = Alignment(indent=1)
    else:
        r[4].alignment = Alignment(indent=2)

for i, w in enumerate([12, 6, 40, 46, 44, 15, 12], 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
sh.freeze_panes = "A2"
sh.auto_filter.ref = f"A1:G{sh.max_row}"
wb.save(OUT)

n1 = sum(1 for x in rows if x[1] == 1)
n2 = sum(1 for x in rows if x[1] == 2)
n3 = sum(1 for x in rows if x[1] == 3)
print(f"{OUT}")
print(f"L1={n1}  L2={n2}  L3={n3}  |  inspeksi total {sum(insp.values())}  |  units {sum(len(v) for v in units.values())}")
