#!/usr/bin/env python3
"""Phase 1 of the canonical-PBS plan: build PBS_MASTER_CANONICAL.xlsx and a DRAFT
data/pbs-crosswalk.csv from PBS_MASTER_INPUT.xlsx (systems 1-6) + the recorded
inspections on the data branch (systems 7-16 and every field variant).

Nothing here touches the data branch. Outputs are for human review; the backfill
(Phase 2) consumes the reviewed crosswalk.

Usage:
  git --work-tree=/tmp/dco checkout origin/data -- data/inspections
  git reset -q -- data/inspections
  python3 scripts/build-canonical-pbs.py [/tmp/dco/data/inspections] [~/Downloads/PBS_MASTER_INPUT.xlsx]
"""
import csv
import json
import glob
import os
import re
import sys
import collections
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
INSPECTIONS = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/dco/data/inspections")
MASTER_IN = Path(sys.argv[2]).expanduser() if len(sys.argv) > 2 else Path.home() / "Downloads" / "PBS_MASTER_INPUT.xlsx"
MASTER_OUT = REPO / "PBS_MASTER_CANONICAL.xlsx"
CROSSWALK_OUT = REPO / "data" / "pbs-crosswalk.csv"

# ---------------------------------------------------------------- review knobs

# Area/Sistem (Level-1) names for systems not in PBS_MASTER_INPUT.xlsx.
SYS_NAME = {
    "7":  "Gas Handling Tanur Reverberatory",
    "8":  "Area Kompresor / Udara Bertekanan",
    "9":  "Gas Handling dan Utilitas Penunjang (Registrasi Lapangan)",
    "11": "Area dan Bangunan Umum",
    "12": "Bangunan dan Gedung Pabrik",
    "13": "Bangunan Workshop dan Rumah Peralatan",
    "14": "Struktur dan Bangunan Tanur / Pemurnian",
    "16": "Area Genset dan Power House",
}
# Sub-sistem (Level-2) name overrides for SYSTEM.MID groups that are a mixed bag
# in the recorded data (else the dominant recorded subsistem is used).
MID_NAME = {
    "9.1": "Peralatan Penunjang (Registrasi Lapangan)",
    "9.3": "Kompresor dan Hydrant",
}
# Whole-code merges: every record at the key PBS is reclassified to the value PBS.
MERGE = {
    "1.3.3":  "1.5.2", "1.3.6": "1.5.2", "1.3.9": "1.5.2",
    "1.3.12": "1.5.2", "1.3.15": "1.5.2",   # Cubicle 20 kV -> Sistem Daya Tanur
    "9.2.4":  "9.2.3",                       # Exhaust Fan (2nd code for same asset)
    "9.2.2":  "9.2.1",                       # ID Fan (per-tanur units of 9.2.1)
    "9.1.5":  "9.1.4",                       # Pneumatic (dup leaf of 9.1.4)
    "9.1.12": "12.1.4",                      # Area Bahan Bakar -> same as 12.1.4
}
# Name-aware merges: only records at (PBS, name-contains) move.
MERGE_BY_NAME = {
    # keys are compared against the typo-folded recorded name
    ("9.1.8",  "gedung crystallizer"): "12.1.14",  # building -> 12.1.14 (Motor Gearbox stays at 9.1.8)
    ("12.1.5", "cooling tower"):       "2.6.2",    # Cooling Tower -> master 2.6.2 (Gedung Komposisi stays at 12.1.5)
}
POSSIBLE_DUP = {}
# namaAsetCanonical overrides that WIN over the master (reviewer decisions).
NAME_OVERRIDE = {
    "2.2.1":  "Pot Crystallizer",              # field "POT 1/2" kept, not master "Screw atau Blitz Crystallizer"
    "12.1.5": "Gedung Komposisi",              # what's left at 12.1.5 after Cooling Tower moved to 2.6.2
    "12.1.4": "Area Bahan Bakar / Gedung HTS MS",  # 12.1.4 has always held both; 9.1.12 merged in
}
# canonical codes whose recorded-vs-master name conflict is DECIDED - suppress the flag.
RESOLVED = {
    "2.4.1", "2.4.2", "2.4.3",        # Burner (master) is right; "Superheater" -> unitLabel
    "2.2.1",                          # -> Pot Crystallizer
    "2.5.5",                          # Bag House Refining (master) is right
    "3.1.2", "3.1.8", "3.2.1", "4.3.1", "2.1.5", "9.1.2",  # master name kept, brand/spec -> unitLabel
    "12.1.4",                         # combined name, 9.1.12 merged in
}
# Canonical Level-3 names for PBS codes NOT in the master (7-16 + out-of-master
# 1-6). If a code is absent here, the cleanest recorded variant is used as-is.
CANON_NAME = {
    "2.5.16": "Cyclone - Pemurnian",
    "7.1.1": "Skin Cooler - DCS Tanur Reverb",
    "7.1.2": "Bag Filter - DCS Tanur Reverb",
    "7.1.3": "ID Fan - DCS Tanur Reverb",
    "7.1.4": "Cerobong - DCS Tanur Reverb",
    "8.1.2": "Distribution Board",
    "9.1.1": "Water / Hydrant",
    "9.1.2": "Water Circulation",
    "9.1.3": "Trafo Step Up 20 kV",
    "9.1.4": "Pneumatic System",
    "9.1.6": "Motor Screw Hopper Elektroda Tanur",
    "9.1.7": "Motor Pump Hydrant",
    "9.1.8": "Motor Gearbox",
    "9.1.9": "Ruang Trafo",
    "9.1.10": "Diesel Pump",
    "9.1.11": "Gudang Kapur",
    "9.1.12": "Area Bahan Bakar",
    "9.1.13": "Gudang LB3",
    "9.2.1": "ID Fan",
    "9.2.3": "Exhaust Fan",
    "9.3.1": "Jockey Pump",
    "9.3.2": "Mesin Compressor Sullair",
    "9.4.1": "Water Instalasi",
    "9.9.1": "Panel Furnace Tegangan Rendah",
    "11.1.1": "Pos Jaga / Area Genset",
    "11.1.2": "Area Peleburan / Tanur",
    "12.1.1": "Blower Vacuum",
    "12.1.2": "Bangunan Crusher",
    "12.1.3": "Gedung Antrasit, Workshop, dan Gudang Pasir",
    "12.1.4": "Gedung HTS MS",
    "12.1.5": "Cooling Tower",
    "12.1.6": "Gedung Musholla",
    "12.1.7": "Gedung Office",
    "12.1.8": "Gedung Rotary Kiln & Flame Oven",
    "12.1.9": "Gedung Timbangan",
    "12.1.10": "Gudang Pasir",
    "12.1.11": "Gudang Sparepart",
    "12.1.12": "Laboratorium",
    "12.1.13": "Gedung Produksi",
    "12.1.14": "Gedung Crystallizer",
    "12.1.15": "Gedung Ruang Pompa",
    "13.1.1": "Workshop Teknik",
    "13.1.2": "Rumah Parkir Alat Berat",
    "13.1.3": "Workshop Tanur Listrik",
    "13.1.4": "Rumah Kompresor",
    "13.1.5": "Rumah Cooling System",
    "13.1.6": "Workshop Listrik",
    "13.1.8": "Rumah Baghouse",
    "14.1.1": "Rumah Mainflue",
    "14.1.2": "Struktur Fixed Tanur",
    "14.1.3": "Struktur Fixed Pemurnian",
    "14.1.4": "Gedung Refining",
    "16.1.1": "Panel Battery",
}

# typo folding is for MATCHING/units only - never rewrites a canonical name
TYPOS = [
    (r"elecroda", "elektroda"), (r"hoistcrane", "hoist crane"),
    (r"kubikel", "cubicle"), (r"kapaistas", "kapasitas"),
    (r"kontruksi", "konstruksi"), (r"pumpt", "pump"), (r"furnance", "furnace"),
    (r"crytalizer", "crystallizer"), (r"battrey", "battery"),
    (r"endtruck", "end truck"), (r"conveyour", "conveyor"),
]

# --------------------------------------------------------------- helpers

def fold(s):
    """lowercase + typo-fold + squeeze whitespace, for comparison only."""
    out = s.lower()
    for pat, rep in TYPOS:
        out = re.sub(pat, rep, out)
    return re.sub(r"\s+", " ", out).strip(" -/,")

norm = fold

def pbskey(code):
    return [int(x) for x in str(code).split(".")]

NOISE = {"kapasitas", "kapaistas", "unit"}

def _toks(s):
    return [t for t in re.split(r"[\s/–-]+", fold(s)) if t]

def _nicecase(t):
    return t.upper() if len(t) == 1 else (t[:1].upper() + t[1:] if t.isalpha() else t)

def unit_of(raw_name, canon_name, merged):
    """Which physical unit the record is about. Merged row -> keep the whole
    original name (clearest for review). Otherwise = folded tokens of raw minus
    any token in the canonical name or NOISE. '' when nothing is left."""
    if fold(raw_name) == fold(canon_name):
        return ""
    if merged:
        return raw_name.strip()
    drop = set(_toks(canon_name)) | NOISE
    keep = [t for t in _toks(raw_name) if t not in drop]
    return " ".join(_nicecase(t) for t in keep)

def words_overlap(a, b):
    return bool(set(_toks(a)) & set(_toks(b)))

# --------------------------------------------------------------- load master 1-6

wb = openpyxl.load_workbook(MASTER_IN, data_only=True)
ws = wb.worksheets[0]
m_l1, m_l2, m_l3 = {}, {}, {}          # "1"->name , "1.2"->name , "1.2.3"->(name, sub, sysname)
cur_sys = cur_sub = None
cur_sysname = cur_subname = None
for kode, lvl, jenis, komp in list(ws.iter_rows(values_only=True))[2:]:
    if lvl == 1:
        cur_sys = str(kode); cur_sysname = komp; m_l1[cur_sys] = komp
    elif lvl == 2:
        cur_sub = str(kode); cur_subname = komp; m_l2[cur_sub] = komp
    elif lvl == 3:
        m_l3[str(kode)] = (komp, cur_subname, cur_sysname)

# --------------------------------------------------------------- load records

pairs = collections.Counter()                       # (pbs_raw, nama_raw) -> n
mid_sub = collections.defaultdict(collections.Counter)
for p in glob.glob(str(INSPECTIONS / "*.json")):
    d = json.load(open(p))
    pbs = str(d.get("nomorPBS", "")).strip()
    if pbs.count(".") != 2:
        continue
    nama = str(d.get("namaAset", "")).strip()
    pairs[(pbs, nama)] += 1
    mid = ".".join(pbs.split(".")[:2])
    mid_sub[mid][str(d.get("subsistem", "")).strip()] += 1

def canonical_pbs(pbs, nama):
    key = (pbs, norm(nama))
    for (mp, sub), tgt in MERGE_BY_NAME.items():
        if mp == pbs and sub in norm(nama):
            return tgt
    return MERGE.get(pbs, pbs)

def sub_name_for(mid):
    if mid in m_l2:
        return m_l2[mid]
    if mid in MID_NAME:
        return MID_NAME[mid]
    c = mid_sub.get(mid)
    return (c.most_common(1)[0][0] if c and c.most_common(1)[0][0] else "(tidak tercatat)")

def sys_name_for(sysn):
    return m_l1.get(sysn) or SYS_NAME.get(sysn) or f"Sistem {sysn}"

def canon_name_for(cpbs):
    if cpbs in NAME_OVERRIDE:
        return NAME_OVERRIDE[cpbs]
    if cpbs in m_l3:
        return m_l3[cpbs][0]
    if cpbs in CANON_NAME:
        return CANON_NAME[cpbs]
    # fallback: cleanest recorded variant at this pbs (shortest)
    cands = [n.strip() for (p, n), _ in pairs.items() if canonical_pbs(p, n) == cpbs and n]
    return min(cands, key=len) if cands else f"PBS {cpbs}"

# --------------------------------------------------------------- crosswalk rows

cw = []
for (pbs, nama), n in sorted(pairs.items(), key=lambda kv: (pbskey(kv[0][0]), kv[0][1])):
    cpbs = canonical_pbs(pbs, nama)
    merged = cpbs != pbs
    mid = ".".join(cpbs.split(".")[:2])
    cname = canon_name_for(cpbs)
    sub = m_l3[cpbs][1] if cpbs in m_l3 else sub_name_for(mid)
    unit = unit_of(nama, cname, merged)
    notes = []
    if merged:
        notes.append(f"merged from {pbs}")
    if cpbs not in m_l3 and cpbs not in CANON_NAME:
        notes.append("canon name = auto (cleanest variant) - REVIEW")
    if cpbs in POSSIBLE_DUP:
        notes.append(POSSIBLE_DUP[cpbs])
    if not merged and nama and cpbs not in RESOLVED and not words_overlap(nama, cname):
        notes.append(f"recorded name unlike canonical ('{cname}') - confirm")
    cw.append({
        "recorded_nomorPBS": pbs,
        "recorded_namaAset": nama,
        "records": n,
        "pbsCanonical": cpbs,
        "subsistemCanonical": sub,
        "namaAsetCanonical": cname,
        "unitLabel": unit,
        "note": "; ".join(notes),
    })

cw.sort(key=lambda r: (pbskey(r["pbsCanonical"]), r["namaAsetCanonical"], r["recorded_namaAset"]))
with open(CROSSWALK_OUT, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=list(cw[0].keys()))
    w.writeheader()
    w.writerows(cw)

# --------------------------------------------------------------- canonical master xlsx

# canonical master = the FULL master systems 1-6 (every L3, even uninspected, so
# the regenerated catalog stays complete) + every canonical code the recorded
# data introduces for systems 7-16 / out-of-master.
cinfo = {}
for r in cw:
    cinfo.setdefault(r["pbsCanonical"], r["namaAsetCanonical"])

canon_codes = set(m_l3) | {r["pbsCanonical"] for r in cw}
canon_codes = sorted(canon_codes, key=pbskey)

out = openpyxl.Workbook()
sh = out.active
sh.title = "PBS Master"
sh.append(["PRODUCT BREAKDOWN STRUCTURE (PBS) - ASET SMELTER (CANONICAL)", None, None, None])
sh.append(["Kode PBS", "Level", "Jenis", "Komponen"])

emitted_sys, emitted_mid = set(), set()
for cpbs in canon_codes:
    sysn = cpbs.split(".")[0]
    mid = ".".join(cpbs.split(".")[:2])
    if sysn not in emitted_sys:
        sh.append([sysn, 1, "Sistem / Area", sys_name_for(sysn)])
        emitted_sys.add(sysn)
    if mid not in emitted_mid:
        sh.append([mid, 2, "Sub-sistem", sub_name_for(mid)])
        emitted_mid.add(mid)
    sh.append([cpbs, 3, "Komponen / Peralatan", canon_name_for(cpbs)])

for i, wdt in enumerate([12, 7, 22, 50], 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
out.save(MASTER_OUT)

# --------------------------------------------------------------- summary

recorded_canon = {r["pbsCanonical"] for r in cw}
print(f"crosswalk rows        : {len(cw)}  -> {CROSSWALK_OUT}")
print(f"  merged rows         : {sum(1 for r in cw if 'merged' in r['note'])}")
print(f"  name-mismatch flags : {sum(1 for r in cw if 'confirm' in r['note'])}  (REVIEW)")
print(f"  rows with any note  : {sum(1 for r in cw if r['note'])}")
print(f"records covered       : {sum(r['records'] for r in cw)}")
print(f"recorded PBS {len({r['recorded_nomorPBS'] for r in cw})}  ->  canonical PBS {len(recorded_canon)}"
      f"  (merged away {len({r['recorded_nomorPBS'] for r in cw}) - len(recorded_canon)})")
print(f"canonical master xlsx : {len(canon_codes)} L3 codes "
      f"({len(m_l3)} full master sys 1-6 + {len(canon_codes) - len(m_l3)} recorded sys 7-16)  -> {MASTER_OUT}")
